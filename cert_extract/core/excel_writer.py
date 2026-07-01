"""Excel 导出。"""

from __future__ import annotations

from pathlib import Path
from typing import Callable, List, Sequence

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from cert_extract.core.models import ExtractRecord
from cert_extract.core.parsers import format_current_certificate_display, normalize_cert_display

ROW_FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ROW_FILL_PEACH = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
COL_FILL_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_FONT = Font(bold=True)
HEADER_FONT_RED = Font(bold=True, color="FF0000")
RED_FONT = Font(color="FF0000")
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
COLUMN_WIDTHS = [6, 18, 52, 38, 38, 42]


def write_excel(
    records: Sequence[ExtractRecord],
    output_path: Path,
    headers: Sequence[str],
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
        if col_idx == 5:
            cell.font = HEADER_FONT_RED
            cell.fill = ROW_FILL_GREEN
        elif col_idx == 6:
            cell.font = HEADER_FONT
            cell.fill = COL_FILL_YELLOW
        else:
            cell.font = HEADER_FONT
            cell.fill = ROW_FILL_GREEN

    for row_idx, record in enumerate(records, 1):
        excel_row = row_idx + 1
        values = [
            row_idx,
            record.rights_holder,
            record.location,
            normalize_cert_display(record.property_right_certificate_no),
            normalize_cert_display(record.original_property_certificate_no),
            format_current_certificate_display(record.property_certificate_no),
        ]
        row_fill = ROW_FILL_GREEN if row_idx % 2 == 1 else ROW_FILL_PEACH
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.alignment = CENTER_ALIGN
            cell.border = THIN_BORDER
            if col_idx == 6:
                cell.fill = COL_FILL_YELLOW
                cell.font = RED_FONT
            else:
                cell.fill = row_fill
                if col_idx == 5:
                    cell.font = RED_FONT

    last_row = len(records) + 1
    ws.auto_filter.ref = f"A1:F{last_row}"
    for col_idx, width in enumerate(COLUMN_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
